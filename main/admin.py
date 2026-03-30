# main/admin.py

from django.conf import settings
from django.contrib import admin, messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Max
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, HttpResponseForbidden
from django.shortcuts import render, redirect
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html
from django import forms

# ========== DJANGO THIRD-PARTY ==========
from modeltranslation.admin import TranslationTabularInline, TranslationStackedInline, TabbedTranslationAdmin, TranslationAdmin
from reversion.admin import VersionAdmin
from reversion.models import Version
from openpyxl.styles import Font, PatternFill, Alignment

# ========== PYTHON STANDARD LIBRARY ==========
import json
import logging
import openpyxl
import os
from datetime import datetime, timedelta
from urllib.parse import unquote

# ========== Р В РІР‚С”Р В РЎвЂєР В РЎв„ўР В РЎвЂ™Р В РІР‚С”Р В Р’В¬Р В РЎСљР В Р’В«Р В РІР‚Сћ Р В Р В РЎС™Р В РЎСџР В РЎвЂєР В Р’В Р В РЎС›Р В Р’В« ==========
from .models import *
from main.services.amocrm.token_manager import TokenManager
logger = logging.getLogger('django')

# ========== Р В РЎСљР В РЎвЂ™Р В Р Р‹Р В РЎС›Р В Р’В Р В РЎвЂєР В РІвЂћСћР В РЎв„ўР В  Р В РЎвЂ™Р В РІР‚СњР В РЎС™Р В Р В РЎСљР В РЎв„ўР В  ==========
admin.site.site_header = "Р В РЎСџР В Р’В°Р В Р вЂ¦Р В Р’ВµР В Р’В»Р РЋР Р‰ Р РЋРЎвЂњР В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ Autoliga"
admin.site.site_title = "Autoliga Admin"
admin.site.index_title = "Р В Р в‚¬Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ Р РЋР С“Р В Р’В°Р В РІвЂћвЂ“Р РЋРІР‚С™Р В Р’В°Р В РЎР В РЎвЂ Autoliga"

# ============ Р В РІР‚Р В РЎвЂ™Р В РІР‚вЂќР В РЎвЂєР В РІР‚в„ўР В Р’В«Р В РІР‚Сћ Р В РЎС™Р В Р В РЎв„ўР В Р Р‹Р В Р В РЎСљР В Р’В« ============

class ContentAdminMixin:
    """Р В РЎС™Р В РЎвЂР В РЎвЂќР РЋР С“Р В РЎвЂР В Р вЂ¦ Р В РўвЂР В Р’В»Р РЋР РЏ Р В РЎвЂќР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р РЋРІР‚С™-Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р В РЎвЂўР В Р вЂ """
    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        

        if request.user.groups.filter(
            name__in=['Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–', 'Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р РЋРІР‚С™-Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–', 'Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р РЋРІР‚С™ UZ']
        ).exists():
            return True
        

        content_models = [
            'product', 'vacancy',
            'featureicon',
        ]
        for model in content_models:
            if request.user.has_perm(f'main.view_{model}'):
                return True
        
        return False
    
    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(
            name__in=['Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–', 'Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р РЋРІР‚С™-Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–', 'Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р РЋРІР‚С™ UZ']
        ).exists():
            return True
        

        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.change_{model_name}')
    
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(name='Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–').exists():
            return True

        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.delete_{model_name}')

class LeadManagerMixin:
    """Р В РЎС™Р В РЎвЂР В РЎвЂќР РЋР С“Р В РЎвЂР В Р вЂ¦ Р В РўвЂР В Р’В»Р РЋР РЏ Р В Р’В»Р В РЎвЂР В РўвЂ-Р В РЎР В Р’ВµР В Р вЂ¦Р В Р’ВµР В РўвЂР В Р’В¶Р В Р’ВµР РЋР вЂљР В РЎвЂўР В Р вЂ """
    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        

        if request.user.groups.filter(
            name__in=['Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–', 'Р В РІР‚С”Р В РЎвЂР В РўвЂ-Р В РЎР В Р’ВµР В Р вЂ¦Р В Р’ВµР В РўвЂР В Р’В¶Р В Р’ВµР РЋР вЂљР РЋРІР‚в„–', 'Р В РІР‚С”Р В РЎвЂР В РўвЂР РЋРІР‚в„– UZ']
        ).exists():
            return True
        

        lead_models = ['contactform', 'jobapplication', 'becomeadealerapplication']
        for model in lead_models:
            if request.user.has_perm(f'main.view_{model}'):
                return True
        
        return False
    
    def has_add_permission(self, request):
        return False 
    
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(name='Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–').exists():
            return True
        

        model_name = self.model._meta.model_name
        return request.user.has_perm(f'main.delete_{model_name}')
        
class AmoCRMAdminMixin:
    """Р В РЎС™Р В РЎвЂР В РЎвЂќР РЋР С“Р В РЎвЂР В Р вЂ¦ Р В РўвЂР В Р’В»Р РЋР РЏ Р РЋРЎвЂњР В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ amoCRM"""
    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        
        if request.user.groups.filter(name='Р В РІР‚СљР В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В°Р В РўвЂР В РЎР В РЎвЂР В Р вЂ¦Р РЋРІР‚в„–').exists():
            return True
        
        return False
    
    def has_view_permission(self, request, obj=None):
        return self.has_module_permission(request)
    
    def has_change_permission(self, request, obj=None):
        return self.has_module_permission(request)
    
    def has_add_permission(self, request):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False

class CustomReversionMixin:
    """Р В РЎС™Р В РЎвЂР В РЎвЂќР РЋР С“Р В РЎвЂР В Р вЂ¦ Р В РўвЂР В Р’В»Р РЋР РЏ Р В РЎвЂќР В Р’В°Р РЋР С“Р РЋРІР‚С™Р В РЎвЂўР В РЎР В Р вЂ¦Р В РЎвЂўР В РЎвЂ“Р В РЎвЂў Р РЋРІвЂљВ¬Р В Р’В°Р В Р’В±Р В Р’В»Р В РЎвЂўР В Р вЂ¦Р В Р’В° Р В Р вЂ Р В РЎвЂўР РЋР С“Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ"""
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'recover/',
                self.admin_site.admin_view(self.custom_recover_list_view),
                name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_recoverlist'
            ),
            path(
                'recover/<int:version_id>/', 
                self.admin_site.admin_view(self.recover_view),  
                name=f'{self.model._meta.app_label}_{self.model._meta.model_name}_recover'
            ),
        ]
        return custom_urls + urls
    
    def custom_recover_list_view(self, request):
        
        opts = self.model._meta
        deleted_versions = Version.objects.get_deleted(self.model)
        
        seen_objects = {}
        version_list_with_preview = []
        
        for version in deleted_versions.order_by('-revision__date_created'):
            obj_repr = version.object_repr
            if obj_repr not in seen_objects:
                seen_objects[obj_repr] = True
                
                preview_url = None
                try:
                    field_dict = version.field_dict
                    for field_name in ['preview_image', 'main_image', 'card_image', 'logo']:
                        if field_name in field_dict and field_dict[field_name]:
                            preview_url = f"{settings.MEDIA_URL}{field_dict[field_name]}"
                            break
                except Exception:
                    pass
                
                version_list_with_preview.append({
                    'version': version,
                    'preview_url': preview_url,
                    'object_repr': obj_repr,
                    'date': version.revision.date_created
                })
        
        context = {
            **self.admin_site.each_context(request),
            'opts': opts,
            'version_list': version_list_with_preview,
            'title': f'Р В РІР‚в„ўР В РЎвЂўР РЋР С“Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ: {opts.verbose_name_plural}',
            'has_view_permission': self.has_view_permission(request),
        }
        
        return render(request, 'admin/reversion/recover_list.html', context)
    
    def recover_view(self, request, version_id):
        
        opts = self.model._meta
        
        try:
            version = Version.objects.get(pk=version_id)
            
            if version.content_type.model_class() != self.model:
                raise Version.DoesNotExist
            
            version.revision.revert()
            
            messages.success(request, f'Р Р†РЎС™РІР‚В¦ Р В РЎвЂєР В Р’В±Р РЋР вЂ°Р В Р’ВµР В РЎвЂќР РЋРІР‚С™ "{version.object_repr}" Р РЋРЎвЂњР РЋР С“Р В РЎвЂ”Р В Р’ВµР РЋРІвЂљВ¬Р В Р вЂ¦Р В РЎвЂў Р В Р вЂ Р В РЎвЂўР РЋР С“Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦!')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_changelist')
            
        except Version.DoesNotExist:
            messages.error(request, 'Р Р†РЎСљР Р‰ Р В РІР‚в„ўР В Р’ВµР РЋР вЂљР РЋР С“Р В РЎвЂР РЋР РЏ Р В Р вЂ¦Р В Р’Вµ Р В Р вЂ¦Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР В Р вЂ¦Р В Р’В° Р В РЎвЂР В Р’В»Р В РЎвЂ Р РЋРЎвЂњР В Р’В¶Р В Р’Вµ Р В Р вЂ Р В РЎвЂўР РЋР С“Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В Р’В°')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')
        except Exception as e:
            messages.error(request, f'Р Р†РЎСљР Р‰ Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р В Р вЂ Р В РЎвЂўР РЋР С“Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ: {str(e)}')
            return redirect(f'admin:{opts.app_label}_{opts.model_name}_recoverlist')

# ============ Р В РЎСљР В РЎвЂєР В РІР‚в„ўР В РЎвЂєР В Р Р‹Р В РЎС›Р В  ============

class NewsBlockInline(TranslationStackedInline): 
    model = NewsBlock
    extra = 1
    fields = ('block_type', 'title', 'text', 'image', 'youtube_url', 'video_file', 'order')

    class Media:
        js = ('js/admin/news_block_dynamic.js',) 
        css = {
            'all': ('css/news_block_custom.css',)  
            
        }

@admin.register(News)
class NewsAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['preview_image_tag', 'title', 'author', 'is_active', 'order', 'created_at', 'action_buttons']
    list_editable = ['is_active', 'order']
    list_filter = ['is_active', 'created_at']
    search_fields = ['title', 'desc']
    readonly_fields = ['preview_image_tag', 'author_photo_tag', 'slug', 'updated_at']
    prepopulated_fields = {}
    inlines = [NewsBlockInline]
    history_latest_first = True
    
    fieldsets = (
        ('Р В РЎвЂєР РЋР С“Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {
            'fields': ('title', 'slug', 'created_at', 'is_active', 'order'),
        }),
        ('Р В РЎв„ўР В Р’В°Р РЋР вЂљР РЋРІР‚С™Р В РЎвЂўР РЋРІР‚РЋР В РЎвЂќР В Р’В° Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂўР РЋР С“Р РЋРІР‚С™Р В РЎвЂ', {
            'fields': ('desc', 'preview_image', 'preview_image_tag'),
        }),
        ('Р В РЎвЂ™Р В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљ', {
            'fields': ('author', 'author_photo', 'author_photo_tag')
        }),
        ('Р В РЎС›Р В Р’ВµР РЋРІР‚В¦Р В Р вЂ¦Р В РЎвЂР РЋРІР‚РЋР В Р’ВµР РЋР С“Р В РЎвЂќР В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {
            'fields': ('updated_at',),
            'classes': ('collapse',)
        }),
    )

    def preview_image_tag(self, obj):
        if obj.preview_image:
            return format_html('<img src="{}" width="100" style="border-radius:8px;"/>', obj.preview_image.url)
        return "Р Р†Р вЂљРІР‚Сњ"
    preview_image_tag.short_description = "Р В РЎСџР РЋР вЂљР В Р’ВµР В Р вЂ Р РЋР Р‰Р РЋР вЂ№"

    def author_photo_tag(self, obj):
        if obj.author_photo:
            return format_html('<img src="{}" width="50" style="border-radius:50%;">', obj.author_photo.url)
        return "Р Р†Р вЂљРІР‚Сњ"
    author_photo_tag.short_description = "Р В Р’В¤Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў Р В Р’В°Р В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР В Р’В°"
    
    def action_buttons(self, obj):
        return format_html('''
            <div style="display: flex; gap: 8px;">
                <a href="{}" title="Р В Р’В Р В Р’ВµР В РўвЂР В Р’В°Р В РЎвЂќР РЋРІР‚С™Р В РЎвЂР РЋР вЂљР В РЎвЂўР В Р вЂ Р В Р’В°Р РЋРІР‚С™Р РЋР Р‰">
                    <img src="/static/media/icon-adminpanel/pencil.png" width="24" height="24">
                </a>
                <a href="/news/{}/" title="Р В РЎСџР РЋР вЂљР В РЎвЂўР РЋР С“Р В РЎР В РЎвЂўР РЋРІР‚С™Р РЋР вЂљ" target="_blank">
                    <img src="/static/media/icon-adminpanel/eyes.png" width="24" height="24">
                </a>
                <a href="{}" title="Р В Р в‚¬Р В РўвЂР В Р’В°Р В Р’В»Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰" onclick="return confirm('Р В Р в‚¬Р В РўвЂР В Р’В°Р В Р’В»Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰?')">
                    <img src="/static/media/icon-adminpanel/recycle-bin.png" width="24" height="24">
                </a>
            </div>
        ''', f'/admin/main/news/{obj.id}/change/', obj.slug, f'/admin/main/news/{obj.id}/delete/')
    action_buttons.short_description = "Р В РІР‚СњР В Р’ВµР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р вЂ Р В РЎвЂР РЋР РЏ"
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        deleted_count = Version.objects.get_deleted(self.model).count()
        if deleted_count > 0:
            extra_context['show_recover_button'] = True
            extra_context['deleted_count'] = deleted_count
        return super().changelist_view(request, extra_context)
    
# ============ Р В РІР‚вЂќР В РЎвЂ™Р В Р вЂЎР В РІР‚в„ўР В РЎв„ўР В  ============

@admin.register(ContactForm)
class ContactFormAdmin(LeadManagerMixin, admin.ModelAdmin):
    change_list_template = 'main/contactform/change_list.html'
    preserve_filters = True
    

    list_select_related = ['manager']
    list_per_page = 50
    show_full_result_count = False
    
    list_display = [
        'name', 'phone', 'product_display', 'region', 
        'priority', 'status', 'amocrm_badge', 
        'manager', 'created_at', 'action_buttons'
    ]
    list_editable = ['priority', 'status', 'manager']  
    list_filter = ['status', 'priority', 'region']
    search_fields = ['name', 'phone', 'amocrm_lead_id']
    readonly_fields = ['created_at', 'amocrm_sent_at', 'amocrm_lead_link']
    autocomplete_fields = ['manager']
    actions = ['retry_failed_leads', 'export_to_excel']

    fieldsets = (
        ('Р В Р В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ Р В РЎвЂў Р В РЎвЂќР В Р’В»Р В РЎвЂР В Р’ВµР В Р вЂ¦Р РЋРІР‚С™Р В Р’Вµ', {
            'fields': ('name', 'phone', 'region', 'message', 'created_at')
        }),
        ('Р В Р в‚¬Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ', {
            'fields': ('status', 'priority', 'manager', 'admin_comment')
        }),
        ('amoCRM', {
            'fields': ('amocrm_status', 'amocrm_lead_link', 'amocrm_sent_at', 'amocrm_error'),
            'classes': ('collapse',)
        }),
    )
    
    class Media:
        css = {'all': ('css/amocrm_modal.css', 'css/contactform_admin.css')}
        js = ('js/amocrm_modal.js', 'js/contactform_admin.js')
    
    # ==================== Р В РЎвЂєР В РЎС›Р В РЎвЂєР В РІР‚Р В Р’В Р В РЎвЂ™Р В РІР‚вЂњР В РІР‚СћР В РЎСљР В Р В РІР‚Сћ ====================
    
    def product_display(self, obj):
        if not obj.product:
            return "Р Р†Р вЂљРІР‚Сњ"
        return format_html(
            '<span style="color:#1976d2;font-weight:600;">{}</span>',
            obj.product[:30]
        )
        return format_html('<span style="color:#999;">Р Р†Р вЂљРІР‚Сњ</span>')
    
    product_display.short_description = "Р В РЎС™Р В РЎвЂўР В РўвЂР В Р’ВµР В Р’В»Р РЋР Р‰"
    product_display.admin_order_field = 'product'
    
    def amocrm_badge(self, obj):
        """Р В РІР‚Р В Р’ВµР В РІвЂћвЂ“Р В РўвЂР В Р’В¶ Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“Р В Р’В° amoCRM"""
        if obj.amocrm_status == 'sent':
            return format_html(
                '<span style="background:#10b981;color:white;padding:5px 12px;border-radius:6px;font-weight:600;font-size:12px;">Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂў</span>'
            )
        elif obj.amocrm_status == 'failed':
            error_text = (obj.amocrm_error or 'Р В РЎСљР В Р’ВµР В РЎвЂР В Р’В·Р В Р вЂ Р В Р’ВµР РЋР С“Р РЋРІР‚С™Р В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В°').replace('"', '&quot;').replace("'", '&#39;')
            return format_html(
                '<span class="amocrm-error-badge" data-error="{}" style="background:#ef4444;color:white;padding:5px 12px;border-radius:6px;font-weight:600;font-size:12px;cursor:pointer;" title="Р В РЎСљР В Р’В°Р В Р’В¶Р В РЎР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РўвЂР В Р’В»Р РЋР РЏ Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР РЋР С“Р В РЎР В РЎвЂўР РЋРІР‚С™Р РЋР вЂљР В Р’В° Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В РЎвЂ">Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В°</span>',
                error_text
            )
        return format_html(
            '<span style="background:#f59e0b;color:white;padding:5px 12px;border-radius:6px;font-weight:600;font-size:12px;">Р В РЎвЂєР В Р’В¶Р В РЎвЂР В РўвЂР В Р’В°Р В Р’ВµР РЋРІР‚С™</span>'
        )

    amocrm_badge.short_description = "amoCRM"
    amocrm_badge.admin_order_field = 'amocrm_status'
    
    def action_buttons(self, obj):
        """Р В РЎв„ўР В Р вЂ¦Р В РЎвЂўР В РЎвЂ”Р В РЎвЂќР В РЎвЂ Р В РўвЂР В Р’ВµР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р вЂ Р В РЎвЂР В РІвЂћвЂ“"""
        view_url = f"https://fawtrucks.amocrm.ru/leads/detail/{obj.amocrm_lead_id}" if obj.amocrm_lead_id else f"/admin/main/contactform/{obj.id}/change/"
        view_title = "Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ  amoCRM" if obj.amocrm_lead_id else "Р В РЎСџР РЋР вЂљР В РЎвЂўР РЋР С“Р В РЎР В РЎвЂўР РЋРІР‚С™Р РЋР вЂљ Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂќР В РЎвЂ"
        
        return format_html('''
            <div style="display:flex;gap:8px;">
                <a href="{}" title="Р В Р’В Р В Р’ВµР В РўвЂР В Р’В°Р В РЎвЂќР РЋРІР‚С™Р В РЎвЂР РЋР вЂљР В РЎвЂўР В Р вЂ Р В Р’В°Р РЋРІР‚С™Р РЋР Р‰" style="padding:6px;border-radius:6px;display:inline-block;transition:transform 0.2s;" onmouseover="this.style.transform='translateY(-2px)'" onmouseout="this.style.transform='translateY(0)'">
                    <img src="/static/media/icon-adminpanel/pencil.png" width="20" height="20">
                </a>
                <a href="{}" title="{}" target="_blank" style="padding:6px;border-radius:6px;display:inline-block;transition:transform 0.2s;" onmouseover="this.style.transform='translateY(-2px)'" onmouseout="this.style.transform='translateY(0)'">
                    <img src="/static/media/icon-adminpanel/eyes.png" width="20" height="20">
                </a>
                <a href="{}" title="Р В Р в‚¬Р В РўвЂР В Р’В°Р В Р’В»Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰" onclick="return confirm('Р В Р в‚¬Р В РўвЂР В Р’В°Р В Р’В»Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂќР РЋРЎвЂњ?')" style="padding:6px;border-radius:6px;display:inline-block;transition:transform 0.2s;" onmouseover="this.style.transform='translateY(-2px)'" onmouseout="this.style.transform='translateY(0)'">
                    <img src="/static/media/icon-adminpanel/recycle-bin.png" width="20" height="20">
                </a>
            </div>
        ''', f'/admin/main/contactform/{obj.id}/change/', view_url, view_title, f'/admin/main/contactform/{obj.id}/delete/')
    
    action_buttons.short_description = "Р В РІР‚СњР В Р’ВµР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р вЂ Р В РЎвЂР РЋР РЏ"
    
    def amocrm_lead_link(self, obj):
        """Р В Р Р‹Р РЋР С“Р РЋРІР‚в„–Р В Р’В»Р В РЎвЂќР В Р’В° Р В Р вЂ¦Р В Р’В° Р В Р’В»Р В РЎвЂР В РўвЂ Р В Р вЂ  amoCRM"""
        if obj.amocrm_lead_id:
            # url = f"https://fawtrucks.amocrm.ru/leads/detail/{obj.amocrm_lead_id}"
            return format_html(
                '<a href="{}" target="_blank" style="color:#3b82f6;font-weight:600;">Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР РЋР вЂљР РЋРІР‚в„–Р РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ  amoCRM (ID: {})</a>',
                url, obj.amocrm_lead_id
            )
        return "Р Р†Р вЂљРІР‚Сњ"
    
    amocrm_lead_link.short_description = "Р В Р Р‹Р РЋР С“Р РЋРІР‚в„–Р В Р’В»Р В РЎвЂќР В Р’В° Р В Р вЂ¦Р В Р’В° Р В Р’В»Р В РЎвЂР В РўвЂ"
    
    # ==================== Р В РІР‚СњР В РІР‚СћР В РІвЂћСћР В Р Р‹Р В РЎС›Р В РІР‚в„ўР В Р В Р вЂЎ ====================
    
    def retry_failed_leads(self, request, queryset):
        """Р В РЎСџР В РЎвЂўР В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂќР В Р’В° Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р вЂ¦Р РЋРІР‚в„–Р РЋРІР‚В¦ Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂўР В РЎвЂќ"""
        logger = logging.getLogger('django')
        
        failed_leads = queryset.filter(amocrm_status='failed')
        
        if not failed_leads.exists():
            self.message_user(request, 'Р В РЎСљР В Р’ВµР РЋРІР‚С™ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р вЂ¦Р РЋРІР‚в„–Р РЋРІР‚В¦ Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂўР В РЎвЂќ Р В РўвЂР В Р’В»Р РЋР РЏ Р В РЎвЂ”Р В РЎвЂўР В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР В Р вЂ¦Р В РЎвЂўР В РІвЂћвЂ“ Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂќР В РЎвЂ', level=messages.WARNING)
            return
        
        success_count = 0
        fail_count = 0
        
        for lead in failed_leads:
            try:
                lead.amocrm_status = 'pending'
                lead.amocrm_error = None
                lead.save()
                
                LeadSender.send_lead(lead)
                lead.refresh_from_db()
                
                if lead.amocrm_status == 'sent':
                    success_count += 1
                else:
                    fail_count += 1
            except Exception as e:
                logger.error(f"Error retrying lead {lead.id}: {str(e)}", exc_info=True)
                fail_count += 1
        
        if success_count > 0:
            self.message_user(request, f'Р В Р в‚¬Р РЋР С“Р В РЎвЂ”Р В Р’ВµР РЋРІвЂљВ¬Р В Р вЂ¦Р В РЎвЂў Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂў: {success_count}', level=messages.SUCCESS)
        if fail_count > 0:
            self.message_user(request, f'Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂќР В РЎвЂ: {fail_count}', level=messages.ERROR)
    
    retry_failed_leads.short_description = 'Р В РЎСџР В РЎвЂўР В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР вЂљР В Р вЂ¦Р В РЎвЂў Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР РЋРІР‚РЋР В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂќР В РЎвЂ'
    
    def export_to_excel(self, request, queryset):
        """Р В Р’В­Р В РЎвЂќР РЋР С“Р В РЎвЂ”Р В РЎвЂўР РЋР вЂљР РЋРІР‚С™ Р В Р вЂ  Excel"""
        logger = logging.getLogger('django')
        
        try:
            if request.POST.get('select_across') == '1':
                queryset = self.get_queryset(request)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Р В РІР‚вЂќР В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂќР В РЎвЂ AutoLiga"
            
            headers = [
                'Р В РЎСљР В РЎвЂўР В РЎР В Р’ВµР РЋР вЂљ', 'Р В Р’В¤Р В Р В РЎвЂє', 'Р В РЎС›Р В Р’ВµР В Р’В»Р В Р’ВµР РЋРІР‚С›Р В РЎвЂўР В Р вЂ¦', 'Р В РЎС™Р В РЎвЂўР В РўвЂР В Р’ВµР В Р’В»Р РЋР Р‰', 'Р В Р’В Р В Р’ВµР В РЎвЂ“Р В РЎвЂР В РЎвЂўР В Р вЂ¦', 'Р В Р Р‹Р В РЎвЂўР В РЎвЂўР В Р’В±Р РЋРІР‚В°Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ', 
                'Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“', 'Р В РЎСџР РЋР вЂљР В РЎвЂР В РЎвЂўР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’ВµР РЋРІР‚С™', 'Р В РЎС™Р В Р’ВµР В Р вЂ¦Р В Р’ВµР В РўвЂР В Р’В¶Р В Р’ВµР РЋР вЂљ', 'Р В РІР‚СњР В Р’В°Р РЋРІР‚С™Р В Р’В°',
                'amoCRM Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“', 'amoCRM ID', 'amoCRM Р В РІР‚СњР В Р’В°Р РЋРІР‚С™Р В Р’В°', 'amoCRM Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В°'
            ]
            ws.append(headers)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for idx, contact in enumerate(queryset, start=1):
                ws.append([
                    idx,
                    contact.name,
                    contact.phone,
                    contact.product[:30] if contact.product else '-',
                    contact.get_region_display(),
                    contact.message[:100] if contact.message else '-',
                    contact.get_status_display(),
                    contact.get_priority_display(),
                    contact.manager.username if contact.manager else '-',
                    contact.created_at.strftime('%d.%m.%Y %H:%M'),
                    contact.get_amocrm_status_display(),
                    contact.amocrm_lead_id or '-',
                    contact.amocrm_sent_at.strftime('%d.%m.%Y %H:%M') if contact.amocrm_sent_at else '-',
                    contact.amocrm_error[:100] if contact.amocrm_error else '-'
                ])
            
            for column in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # response['Content-Disposition'] = f'attachment; filename="faw_uz_contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            
            return response
            
        except Exception as e:
            logger.error(f"Р Р†РЎСљР Р‰ Error exporting to Excel: {str(e)}", exc_info=True)
            self.message_user(request, f'Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р РЋР РЉР В РЎвЂќР РЋР С“Р В РЎвЂ”Р В РЎвЂўР РЋР вЂљР РЋРІР‚С™Р В Р’В°: {str(e)}', level=messages.ERROR)
            return redirect(request.path)
    
    export_to_excel.short_description = 'Р В Р’В­Р В РЎвЂќР РЋР С“Р В РЎвЂ”Р В РЎвЂўР РЋР вЂљР РЋРІР‚С™ Р В Р вЂ  Excel'
    
    # ==================== QUERYSET ====================
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР В Р’В»Р РЋР вЂ№Р РЋРІР‚РЋР В Р’В°Р В Р’ВµР В РЎ Р РЋРЎвЂњР В РЎвЂ”Р РЋР вЂљР В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ Р В РЎР В Р’ВµР В Р вЂ¦Р В Р’ВµР В РўвЂР В Р’В¶Р В Р’ВµР РЋР вЂљР В Р’В°Р В РЎР В РЎвЂ"""
        formfield = super().formfield_for_foreignkey(db_field, request, **kwargs)
        if db_field.name == "manager":
            formfield.widget.can_add_related = False
            formfield.widget.can_change_related = False
            formfield.widget.can_delete_related = False
            formfield.widget.can_view_related = False
        return formfield
    
    def get_queryset(self, request):
        """Р В Р’В¤Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ queryset"""
        qs = super().get_queryset(request)
        
        # Р В РЎСџР В РЎвЂўР В РЎвЂР РЋР С“Р В РЎвЂќ
        if search_query := request.GET.get('q', '').strip():
            qs = qs.filter(
                Q(name__icontains=search_query) | 
                Q(phone__icontains=search_query) | 
                Q(amocrm_lead_id__icontains=search_query)
            )

        if status := request.GET.get('status', '').strip():
            qs = qs.filter(status=status)
        
        if amocrm_status := request.GET.get('amocrm_status', '').strip():
            qs = qs.filter(amocrm_status=amocrm_status)
        
        if priority := request.GET.get('priority', '').strip():
            qs = qs.filter(priority=priority)
        
        if region := request.GET.get('region', '').strip():
            qs = qs.filter(region=region)
        
        if product := request.GET.get('product', '').strip():
            qs = qs.filter(product__icontains=product)
        

        if date_from := request.GET.get('date_from', '').strip():
            try:
                parsed_date = datetime.strptime(date_from, '%Y-%m-%d')
                date_from_aware = timezone.make_aware(
                    parsed_date.replace(hour=0, minute=0, second=0, microsecond=0)
                )
                qs = qs.filter(created_at__gte=date_from_aware)
            except ValueError:
                pass
        
        if date_to := request.GET.get('date_to', '').strip():
            try:
                parsed_date = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_aware = timezone.make_aware(
                    parsed_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                )
                qs = qs.filter(created_at__lte=date_to_aware)
            except ValueError:
                pass
        
        return qs

    def get_changelist(self, request, **kwargs):
        """Р В РЎСџР В Р’ВµР РЋР вЂљР В Р’ВµР В РЎвЂўР В РЎвЂ”Р РЋР вЂљР В Р’ВµР В РўвЂР В Р’ВµР В Р’В»Р РЋР РЏР В Р’ВµР В РЎ ChangeList Р РЋРІР‚РЋР РЋРІР‚С™Р В РЎвЂўР В Р’В±Р РЋРІР‚в„– Р В РЎвЂР В РЎвЂ“Р В Р вЂ¦Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР вЂљР В РЎвЂўР В Р вЂ Р В Р’В°Р РЋРІР‚С™Р РЋР Р‰ date_from/date_to"""
        from django.contrib.admin.views.main import ChangeList
        
        class CustomChangeList(ChangeList):
            def get_filters_params(self, params=None):
                """Р В Р в‚¬Р В Р’В±Р В РЎвЂР РЋР вЂљР В Р’В°Р В Р’ВµР В РЎ date_from Р В РЎвЂ date_to Р В РЎвЂР В Р’В· lookup Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В Р вЂ """
                lookup_params = super().get_filters_params(params)
                

                lookup_params.pop('date_from', None)
                lookup_params.pop('date_to', None)
                
                return lookup_params
        
        return CustomChangeList
    
    def changelist_view(self, request, extra_context=None):
        """Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’ВµР В РЎвЂќР РЋР С“Р РЋРІР‚С™ Р В РўвЂР В Р’В»Р РЋР РЏ Р РЋРІР‚С›Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В Р вЂ """
        extra_context = extra_context or {}
        
        from main.models import REGION_CHOICES
        extra_context['regions'] = REGION_CHOICES
        
        products = ContactForm.objects.exclude(
            product__isnull=True
        ).exclude(
            product=''
        ).values_list('product', flat=True).distinct().order_by('product')
        
        extra_context['products'] = list(products)
        
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:object_id>/quick-update/', self.admin_site.admin_view(self.quick_update_view), name='contactform_quick_update'),
        ]
        return custom_urls + urls

    def quick_update_view(self, request, object_id):
        """AJAX Р В Р’В°Р В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР РЋР С“Р В РЎвЂўР РЋРІР‚В¦Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ Р РЋР С“Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“Р В Р’В°/Р В РЎвЂ”Р РЋР вЂљР В РЎвЂР В РЎвЂўР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’ВµР РЋРІР‚С™Р В Р’В°/Р В РЎР В Р’ВµР В Р вЂ¦Р В Р’ВµР В РўвЂР В Р’В¶Р В Р’ВµР РЋР вЂљР В Р’В°"""
        import json
        from django.http import JsonResponse
        
        if request.method != 'POST':
            return JsonResponse({'error': 'Method not allowed'}, status=405)
        
        try:
            obj = ContactForm.objects.get(pk=object_id)
            data = json.loads(request.body)
            
            if 'status' in data:
                obj.status = data['status']
            if 'priority' in data:
                obj.priority = data['priority']
            if 'manager' in data:
                obj.manager_id = data['manager'] if data['manager'] else None
            
            obj.save()
            
            return JsonResponse({'success': True})
        
        except ContactForm.DoesNotExist:
            return JsonResponse({'error': 'Object not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
# ============ Р В РІР‚в„ўР В РЎвЂ™Р В РЎв„ўР В РЎвЂ™Р В РЎСљР В Р Р‹Р В Р В  ============

class VacancyResponsibilityInline(TranslationStackedInline):
    model = VacancyResponsibility
    extra = 2
    fields = (('title', 'order'), 'text')

class VacancyRequirementInline(TranslationTabularInline):
    model = VacancyRequirement
    extra = 3
    fields = ('text', 'order')

class VacancyConditionInline(TranslationTabularInline):
    model = VacancyCondition
    extra = 3
    fields = ('text', 'order')

class VacancyIdealCandidateInline(TranslationTabularInline):
    model = VacancyIdealCandidate
    extra = 3
    fields = ('text', 'order')

@admin.register(Vacancy)
class VacancyAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['title', 'is_active', 'applications_count', 'order', 'created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['title', 'short_description']
    prepopulated_fields = {'slug': ('title',)}
    readonly_fields = ['created_at', 'updated_at', 'applications_count']
    inlines = [VacancyResponsibilityInline, VacancyRequirementInline, VacancyIdealCandidateInline, VacancyConditionInline]
    history_latest_first = True
    
    fieldsets = (
        ('Р В РЎвЂєР РЋР С“Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {'fields': ('title', 'slug', 'short_description', 'is_active', 'order')}),
        ('Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’В°Р В РЎвЂќР РЋРІР‚С™Р РЋРІР‚в„–', {'fields': ('contact_info',)}),
        ('Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р В РЎвЂР В РЎвЂќР В Р’В°', {'fields': ('applications_count', 'created_at', 'updated_at'), 'classes': ('collapse',)}),
    )
    
    def applications_count(self, obj):
        count = obj.get_applications_count()
        if count > 0:
            return format_html(
                '<a href="/admin/main/jobapplication/?vacancy__id__exact={}" style="color:#007bff;font-weight:bold;"> {} Р В РІР‚вЂќР В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂўР В РЎвЂќ</a>',
                obj.id, count
            )
        return '0 Р В Р’В·Р В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂўР В РЎвЂќ'
    applications_count.short_description = 'Р В РІР‚вЂќР В Р’В°Р РЋР РЏР В Р вЂ Р В РЎвЂќР В РЎвЂ'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        deleted_count = Version.objects.get_deleted(self.model).count()
        if deleted_count > 0:
            extra_context['show_recover_button'] = True
            extra_context['deleted_count'] = deleted_count
        return super().changelist_view(request, extra_context)

@admin.register(JobApplication)
class JobApplicationAdmin(LeadManagerMixin, admin.ModelAdmin):
    list_display = ['vacancy', 'region', 'applicant_name', 'resume_link', 'file_size_display', 'created_at', 'is_processed_badge']
    list_filter = ['is_processed', 'vacancy', 'region', 'created_at']
    search_fields = ['applicant_name', 'applicant_phone', 'applicant_email', 'vacancy__title']
    readonly_fields = ['created_at', 'file_size_display', 'resume_preview']
    date_hierarchy = 'created_at'
    autocomplete_fields = ['vacancy']
    
    fieldsets = (
        ('Р В Р В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {'fields': ('vacancy', 'region', 'created_at')}),
        ('Р В Р’В Р В Р’ВµР В Р’В·Р РЋР вЂ№Р В РЎР В Р’Вµ', {'fields': ('resume', 'file_size_display', 'resume_preview')}),
        ('Р В РЎв„ўР В РЎвЂўР В Р вЂ¦Р РЋРІР‚С™Р В Р’В°Р В РЎвЂќР РЋРІР‚С™Р РЋРІР‚в„–', {'fields': ('applicant_name', 'applicant_phone', 'applicant_email')}),
        ('Р В РЎвЂєР В Р’В±Р РЋР вЂљР В Р’В°Р В Р’В±Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂќР В Р’В°', {'fields': ('is_processed', 'admin_comment')}),
    )
    
    def resume_link(self, obj):
        if obj.resume:
            return format_html('<a href="{}" target="_blank" style="color:#007bff;font-weight:bold;"> Р В Р Р‹Р В РЎвЂќР В Р’В°Р РЋРІР‚РЋР В Р’В°Р РЋРІР‚С™Р РЋР Р‰</a>', obj.resume.url)
        return "Р Р†Р вЂљРІР‚Сњ"
    resume_link.short_description = 'Р В Р’В Р В Р’ВµР В Р’В·Р РЋР вЂ№Р В РЎР В Р’Вµ'
    
    def file_size_display(self, obj):
        size = obj.get_file_size()
        return f"{size} MB" if size else "Р Р†Р вЂљРІР‚Сњ"
    file_size_display.short_description = 'Р В Р’В Р В Р’В°Р В Р’В·Р В РЎР В Р’ВµР РЋР вЂљ'
    
    def resume_preview(self, obj):
        if obj.resume:
            file_ext = obj.resume.name.split('.')[-1].lower()
            if file_ext in ['jpg', 'jpeg', 'png']:
                return format_html('<img src="{}" width="300" style="border-radius:8px;">', obj.resume.url)
            return format_html('<p style="color:#888;"> {}</p>', obj.resume.name)
        return "Р Р†Р вЂљРІР‚Сњ"
    resume_preview.short_description = 'Р В РЎСџР РЋР вЂљР В Р’ВµР В Р вЂ Р РЋР Р‰Р РЋР вЂ№'
    
    def is_processed_badge(self, obj):
        if obj.is_processed:
            return format_html('<span style="color:green;font-weight:bold;"> Р В Р’В Р В Р’В°Р РЋР С“Р РЋР С“Р В РЎР В РЎвЂўР РЋРІР‚С™Р РЋР вЂљР В Р’ВµР В Р вЂ¦Р В РЎвЂў</span>')
        return format_html('<span style="color:orange;font-weight:bold;"> Р В РЎСљР В РЎвЂўР В Р вЂ Р В Р’В°Р РЋР РЏ</span>')
    is_processed_badge.short_description = 'Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“'

# ============ Р В Р В РЎв„ўР В РЎвЂєР В РЎСљР В РЎв„ўР В  ============

@admin.register(FeatureIcon)
class FeatureIconAdmin(ContentAdminMixin, admin.ModelAdmin):
    list_display = ['icon_preview', 'name', 'order']
    list_editable = ['name', 'order']
    search_fields = ['name']
    
    def icon_preview(self, obj):
        if obj.icon:
            return format_html('<img src="{}" width="30" height="30"/>', obj.icon.url)
        return "Р Р†Р вЂљРІР‚Сњ"
    icon_preview.short_description = "Р В РЎСџР РЋР вЂљР В Р’ВµР В Р вЂ Р РЋР Р‰Р РЋР вЂ№"

# ============ Р В РЎСџР В Р’В Р В РЎвЂєР В РІР‚СњР В Р в‚¬Р В РЎв„ўР В РЎС›Р В Р’В« ============

@admin.register(ProductCategory)
class ProductCategoryAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    list_display = ['icon_preview', 'name', 'slug', 'products_count', 'is_active', 'order']
    list_editable = ['is_active', 'order']
    search_fields = ['name', 'slug']
    prepopulated_fields = {'slug': ('name',)}
    readonly_fields = ['products_count', 'created_at']
    
    fieldsets = (
        ('Р В РЎвЂєР РЋР С“Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {
            'fields': ('name', 'slug', 'description', 'is_active', 'order')
        }),
        ('Р В Р В Р’В·Р В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’В°Р В Р’В¶Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ', {
            'fields': ('icon', 'hero_image')
        }),
        ('Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р В РЎвЂР В РЎвЂќР В Р’В°', {
            'fields': ('products_count', 'created_at'),
            'classes': ('collapse',)
        }),
    )
    
    def icon_preview(self, obj):
        if obj.icon:
            return format_html('<img src="{}" width="30" height="30"/>', obj.icon.url)
        return "Р Р†Р вЂљРІР‚Сњ"
    icon_preview.short_description = "Р В Р В РЎвЂќР В РЎвЂўР В Р вЂ¦Р В РЎвЂќР В Р’В°"
    
    def products_count(self, obj):
        count = obj.get_products_count()
        if count > 0:
            return format_html(
                '<a href="/admin/main/product/?category__id__exact={}" style="color:#007bff;font-weight:bold;">{} Р РЋРІР‚С™Р В РЎвЂўР В Р вЂ Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р вЂ </a>',
                obj.id, count
            )
        return '0 Р РЋРІР‚С™Р В РЎвЂўР В Р вЂ Р В Р’В°Р РЋР вЂљР В РЎвЂўР В Р вЂ '
    products_count.short_description = 'Р В РЎСџР РЋР вЂљР В РЎвЂўР В РўвЂР РЋРЎвЂњР В РЎвЂќР РЋРІР‚С™Р В РЎвЂўР В Р вЂ '

# ========== Р В РЎСџР В Р’В Р В РЎвЂєР В РІР‚СњР В Р в‚¬Р В РЎв„ўР В РЎС›Р В Р’В« (Р В РЎвЂєР В РІР‚Р В РЎСљР В РЎвЂєР В РІР‚в„ўР В РІР‚С”Р В РІР‚СћР В РЎСљР В РЎСљР В Р’В«Р В РІвЂћСћ) ==========

# ========== Р В РЎСџР В Р’В Р В РЎвЂєР В РІР‚СњР В Р в‚¬Р В РЎв„ўР В РЎС›Р В Р’В« (Р В РЎвЂєР В РІР‚Р В РЎСљР В РЎвЂєР В РІР‚в„ўР В РІР‚С”Р В РІР‚СћР В РЎСљР В РЎСљР В Р’В«Р В РІвЂћСћ) ==========

class ProductCategoryFilter(admin.SimpleListFilter):
    """Р В Р’В¤Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљ Р В РЎвЂ”Р В РЎвЂў Р В РЎР В Р’В°Р РЋР вЂљР В РЎвЂќР В Р’Вµ Р В Р’В°Р В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР В РЎР В РЎвЂўР В Р’В±Р В РЎвЂР В Р’В»Р РЋР РЏ"""
    title = 'Р В РЎС™Р В Р’В°Р РЋР вЂљР В РЎвЂќР В Р’В° Р В Р’В°Р В Р вЂ Р РЋРІР‚С™Р В РЎвЂўР В РЎР В РЎвЂўР В Р’В±Р В РЎвЂР В Р’В»Р РЋР РЏ'
    parameter_name = 'category_filter'
    
    def lookups(self, request, model_admin):
        categories = ProductCategory.objects.filter(is_active=True).order_by('order', 'name')
        return [(cat.id, cat.name) for cat in categories]
    
    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(category__id=self.value())
        return queryset


class ParameterCategoryFilter(admin.SimpleListFilter):
    """Р В Р’В¤Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљ Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В Р вЂ  Р В РЎвЂ”Р В РЎвЂў Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР В РЎвЂ (Р В РўвЂР В Р’В»Р РЋР РЏ ProductAdmin)"""
    title = 'Р В РЎв„ўР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР РЏ Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР В Р’В°'
    parameter_name = 'param_category'

    def lookups(self, request, model_admin):
        cats = ParameterCategory.objects.filter(is_active=True).order_by('order', 'name')
        return [(cat.id, cat.name) for cat in cats]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(parameters__category__id=self.value()).distinct()
        return queryset


# class ProductParameterInline(admin.TabularInline):
#     """
#     Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР РЋРІР‚в„– Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР РЋРЎвЂњР В РЎвЂќР РЋРІР‚С™Р В Р’В°.
#     Р Р†РЎС™РІР‚В¦ TranslationTabularInline Р РЋР РЉР В РЎР В Р’В°Р РЋР С“ Р Р†Р вЂљРІР‚Сњ category Р РЋР РЉР В Р вЂ¦Р В РўвЂР В РЎвЂ ForeignKey (tarjima Р В РІвЂћвЂ“Р РЋРЎвЂєР СћРІР‚С”).
#     """
#     model = ProductParameter
#     extra = 0
#     fields = ('category', 'text', 'order')
#     verbose_name = "Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљ"
#     verbose_name_plural = "РЎР‚РЎСџРІР‚СљРІР‚в„– Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР РЋРІР‚в„– (Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР вЂ№ Р В РўвЂР В Р’В»Р РЋР РЏ Р РЋРІР‚С›Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР В РЎвЂ)"
#     # category dropdown Р РЋРЎвЂњР РЋРІР‚РЋР РЋРЎвЂњР В Р вЂ¦ autocomplete (ParameterCategoryAdmin Р В РўвЂР В Р’В° search_fields Р В Р’В±Р РЋРЎвЂєР В Р’В»Р В РЎвЂР РЋРІвЂљВ¬Р В РЎвЂ Р В РЎвЂќР В Р’ВµР РЋР вЂљР В Р’В°Р В РЎвЂќ)
#     autocomplete_fields = ['category']


    # class Media:
    #     js = ('js/admin/parameter_filter.js',)
    #     css = {'all': ('css/admin/parameter_filter.css',)}    

class ProductParameterInline(TranslationTabularInline):
    """Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР РЋРІР‚в„– Р РЋР С“ Р РЋРІР‚С›Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР В Р’ВµР В РІвЂћвЂ“ Р В РЎвЂ”Р В РЎвЂў Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР В РЎвЂ"""
    model = ProductParameter
    extra = 0
    fields = ('category', 'text', 'order')
    verbose_name = "Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљ"
    verbose_name_plural = "РЎР‚РЎСџРІР‚СљРІР‚в„– Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР РЋРІР‚в„– (Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋРІР‚С™Р В Р’Вµ Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР вЂ№ Р В РўвЂР В Р’В»Р РЋР РЏ Р РЋРІР‚С›Р В РЎвЂР В Р’В»Р РЋР Р‰Р РЋРІР‚С™Р РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР В РЎвЂ)"
    
    class Media:
        js = ('js/admin/parameter_filter.js',)
        css = {'all': ('css/admin/parameter_filter.css',)}




@admin.register(ParameterCategory)
class ParameterCategoryAdmin(TranslationAdmin):
    """
    Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљ Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР РЏР В Р’В»Р В Р’В°Р РЋР вЂљР В РЎвЂ Р Р†Р вЂљРІР‚Сњ Р РЋРІР‚С™Р РЋРЎвЂєР В Р’В»Р В РЎвЂР СћРІР‚С” CRUD, 3 Р РЋРІР‚С™Р В РЎвЂР В Р’В»Р В РўвЂР В Р’В°.
    """
    list_display = ('name', 'slug', 'order', 'is_active', 'get_parameters_count')
    list_editable = ('order', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('name', 'name_uz', 'name_ru', 'name_en', 'slug')
    prepopulated_fields = {'slug': ('name',)}

    fieldsets = (
        (None, {
            'fields': ('slug', 'order', 'is_active')
        }),
        ("РЎР‚РЎСџРІР‚РЋРЎвЂќРЎР‚РЎСџРІР‚РЋРЎвЂ” O'zbek", {
            'fields': ('name_uz',),
            'classes': ('collapse',),
        }),
        ("РЎР‚РЎСџРІР‚РЋР’В·РЎР‚РЎСџРІР‚РЋРЎвЂќ Р В Р’В Р РЋРЎвЂњР РЋР С“Р РЋР С“Р В РЎвЂќР В РЎвЂР В РІвЂћвЂ“", {
            'fields': ('name_ru',),
            'classes': ('collapse',),
        }),
        ("РЎР‚РЎСџРІР‚РЋР’В¬РЎР‚РЎСџРІР‚РЋР’В§ English", {
            'fields': ('name_en',),
            'classes': ('collapse',),
        }),
    )

    @admin.display(description="Р В РЎСџР В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР В Р’В»Р В Р’В°Р РЋР вЂљ Р РЋР С“Р В РЎвЂўР В Р вЂ¦Р В РЎвЂ")
    def get_parameters_count(self, obj):
        count = obj.parameters.count()
        return f"{count} Р РЋРІР‚С™Р В Р’В°"

class ProductFeatureInline(TranslationTabularInline):
    model = ProductFeature
    extra = 0
    max_num = 8
    fields = ('icon', 'name', 'order')
    verbose_name = "Р В РўС’Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎвЂќР РЋРІР‚С™Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋР С“Р РЋРІР‚С™Р В РЎвЂР В РЎвЂќР В Р’В°"
    verbose_name_plural = "РЎР‚РЎСџРІР‚СњРІвЂћвЂ“ Р В РўС’Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎвЂќР РЋРІР‚С™Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋР С“Р РЋРІР‚С™Р В РЎвЂР В РЎвЂќР В РЎвЂ Р РЋР С“ Р В РЎвЂР В РЎвЂќР В РЎвЂўР В Р вЂ¦Р В РЎвЂќР В Р’В°Р В РЎР В РЎвЂ"


class ProductCardSpecInline(TranslationTabularInline):
    model = ProductCardSpec
    extra = 0
    max_num = 4
    fields = ('icon', 'value', 'order')
    verbose_name = "Р В Р Р‹Р В РЎвЂ”Р В Р’ВµР РЋРІР‚В Р В РЎвЂР РЋРІР‚С›Р В РЎвЂР В РЎвЂќР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ"
    verbose_name_plural = "РЎР‚РЎСџРІР‚СљРІР‚С› Р В РўС’Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎвЂќР РЋРІР‚С™Р В Р’ВµР РЋР вЂљР В РЎвЂР РЋР С“Р РЋРІР‚С™Р В РЎвЂР В РЎвЂќР В РЎвЂ Р В РЎвЂќР В Р’В°Р РЋР вЂљР РЋРІР‚С™Р В РЎвЂўР РЋРІР‚РЋР В РЎвЂќР В РЎвЂ"


class ProductGalleryInline(admin.TabularInline):
    model = ProductGallery
    extra = 1
    fields = ('image', 'order')
    verbose_name = "Р В Р’В¤Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў"
    verbose_name_plural = "РЎР‚РЎСџРІР‚вЂњРЎР С—РЎвЂР РЏ Р В РІР‚СљР В Р’В°Р В Р’В»Р В Р’ВµР РЋР вЂљР В Р’ВµР РЋР РЏ"


@admin.register(Product)
class ProductAdmin(ContentAdminMixin, CustomReversionMixin, VersionAdmin, TabbedTranslationAdmin):
    list_display = ['thumbnail', 'title', 'category_display', 'is_active', 'is_featured', 'slider_order', 'order']
    list_filter = [ProductCategoryFilter, 'is_active', 'is_featured']
    search_fields = ['title', 'slug']
    list_editable = ['is_active', 'is_featured', 'slider_order', 'order']
    prepopulated_fields = {'slug': ('title',)}
    history_latest_first = True
    actions = ['add_to_slider', 'remove_from_slider']
    
    list_per_page = 15
    show_full_result_count = False
    
    fieldsets = (
        ('Р В РЎвЂєР РЋР С“Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {
            'fields': (
                ('title', 'slug'),
                'category',  # Р Р†РЎС™РІР‚В¦ Р В РЎС›Р В Р’ВµР В РЎвЂ”Р В Р’ВµР РЋР вЂљР РЋР Р‰ Р В РЎвЂўР В Р’В±Р РЋРІР‚в„–Р РЋРІР‚РЋР В Р вЂ¦Р РЋРІР‚в„–Р В РІвЂћвЂ“ select
                ('order', 'is_active', 'is_featured'),
                ('main_image', 'card_image')
            )
        }),
        ('Р Р†Р’В­РЎвЂ™ Р В РЎСљР В Р’В°Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В РІвЂћвЂ“Р В РЎвЂќР В РЎвЂ Р В РЎвЂ“Р В Р’В»Р В Р’В°Р В Р вЂ Р В Р вЂ¦Р В РЎвЂўР В РЎвЂ“Р В РЎвЂў Р РЋР С“Р В Р’В»Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР РЋР вЂљР В Р’В°', {
            'classes': ('collapse',),
            'fields': (
                'slider_image',
                ('slider_year', 'slider_order'),
                'slider_price',
                ('slider_power', 'slider_fuel_consumption'),
            ),
        }),
    )
    
    inlines = [ProductParameterInline, ProductFeatureInline, ProductCardSpecInline, ProductGalleryInline]
    
    def thumbnail(self, obj):
        img = obj.card_image or obj.main_image
        if img:
            return format_html(
                '<img src="{}" width="80" height="50" style="object-fit:cover;border-radius:4px;"/>',
                img.url
            )
        return "Р Р†Р вЂљРІР‚Сњ"
    thumbnail.short_description = "Р В Р’В¤Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў"

    def category_display(self, obj):
        """Р Р†РЎС™РІР‚В¦ Р В РЎСљР В РЎвЂєР В РІР‚в„ўР В РЎвЂєР В РІР‚Сћ: Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’В°Р В Р’В¶Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ Р В РЎвЂўР В РўвЂР В Р вЂ¦Р В РЎвЂўР В РІвЂћвЂ“ Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР В РЎвЂ"""
        if obj.category:
            return format_html(
                '<span style="background:#1976d2;color:white;padding:4px 8px;border-radius:4px;font-size:11px;font-weight:600;">{}</span>',
                obj.category.name
            )
        return format_html('<span style="color:#999;">Р В РІР‚Р В Р’ВµР В Р’В· Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР В РЎвЂ</span>')
    category_display.short_description = "Р В РЎв„ўР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР РЋР РЏ"
    
    def add_to_slider(self, request, queryset):
        updated = queryset.update(is_featured=True)
        self.message_user(request, f'Р Р†РЎС™РІР‚В¦ {updated} Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР РЋРЎвЂњР В РЎвЂќР РЋРІР‚С™Р В РЎвЂўР В Р вЂ  Р В РўвЂР В РЎвЂўР В Р’В±Р В Р’В°Р В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂў Р В Р вЂ  Р РЋР С“Р В Р’В»Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР РЋР вЂљ')
    add_to_slider.short_description = 'Р Р†Р’В­РЎвЂ™ Р В РІР‚СњР В РЎвЂўР В Р’В±Р В Р’В°Р В Р вЂ Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ  Р РЋР С“Р В Р’В»Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР РЋР вЂљ'
    
    def remove_from_slider(self, request, queryset):
        updated = queryset.update(is_featured=False)
        self.message_user(request, f'Р Р†РЎСљР Р‰ {updated} Р В РЎвЂ”Р РЋР вЂљР В РЎвЂўР В РўвЂР РЋРЎвЂњР В РЎвЂќР РЋРІР‚С™Р В РЎвЂўР В Р вЂ  Р РЋРЎвЂњР В Р’В±Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В РЎвЂў Р В РЎвЂР В Р’В· Р РЋР С“Р В Р’В»Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР РЋР вЂљР В Р’В°')
    remove_from_slider.short_description = 'Р Р†РЎСљР Р‰ Р В Р в‚¬Р В Р’В±Р РЋР вЂљР В Р’В°Р РЋРІР‚С™Р РЋР Р‰ Р В РЎвЂР В Р’В· Р РЋР С“Р В Р’В»Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР РЋР вЂљР В Р’В°'

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        deleted_count = Version.objects.get_deleted(self.model).count()
        if deleted_count > 0:
            extra_context['show_recover_button'] = True
            extra_context['deleted_count'] = deleted_count
        featured_count = Product.objects.filter(is_featured=True, is_active=True).count()
        extra_context['featured_count'] = featured_count
        extra_context['show_slider_info'] = True
        return super().changelist_view(request, extra_context)
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'api/parameter-suggestions/',
                self.admin_site.admin_view(self.parameter_suggestions_api),
                name='parameter_suggestions_api'
            ),
        ]
        return custom_urls + urls

    def parameter_suggestions_api(self, request):
        """API Р В РўвЂР В Р’В»Р РЋР РЏ Р В РЎвЂ”Р В РЎвЂўР В Р’В»Р РЋРЎвЂњР РЋРІР‚РЋР В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ Р В РЎвЂ”Р В РЎвЂўР В РўвЂР РЋР С“Р В РЎвЂќР В Р’В°Р В Р’В·Р В РЎвЂўР В РЎвЂќ Р В РЎвЂ”Р В Р’В°Р РЋР вЂљР В Р’В°Р В РЎР В Р’ВµР РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В Р вЂ  Р В РЎвЂ”Р В РЎвЂў Р В РЎвЂќР В Р’В°Р РЋРІР‚С™Р В Р’ВµР В РЎвЂ“Р В РЎвЂўР РЋР вЂљР В РЎвЂР В РЎвЂ"""
        category = request.GET.get('category', '')
        
        if not category:
            return JsonResponse({'suggestions': []})
        
        from django.db.models import Count
        
        suggestions = ProductParameter.objects.filter(
            category=category
        ).values('text').annotate(
            usage_count=Count('id')
        ).order_by('-usage_count')[:20]
        
        result = []
        seen = set()
        
        for item in suggestions:
            text = item['text']
            if text and text not in seen:
                seen.add(text)
                result.append({
                    'text': text,
                    'count': item['usage_count']
                })
        
        return JsonResponse({'suggestions': result})

@admin.register(AmoCRMToken)
class AmoCRMTokenAdmin(AmoCRMAdminMixin, admin.ModelAdmin):
    list_display = ['token_status', 'expires_display', 'time_left_display', 'action_buttons']
    
    # ========== Р В РЎвЂєР В РЎС›Р В РЎвЂєР В РІР‚Р В Р’В Р В РЎвЂ™Р В РІР‚вЂњР В РІР‚СћР В РЎСљР В Р В РІР‚Сћ ==========
    def token_status(self, obj):
        """Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“ Р РЋРІР‚С™Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦Р В Р’В°"""
        
        if not obj.access_token:
            return format_html(
                '<span style="background:#dc3545;color:white;padding:6px 12px;border-radius:6px;font-weight:600;">Р В РЎСљР В Р’Вµ Р В Р вЂ¦Р В Р’В°Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В Р’ВµР В Р вЂ¦</span>'
            )
        
        if obj.is_expired():
            return format_html(
                '<span style="background:#ffc107;color:#000;padding:6px 12px;border-radius:6px;font-weight:600;">Р В Р РЋР С“Р РЋРІР‚С™Р В Р’ВµР В РЎвЂќР В Р’В°Р В Р’ВµР РЋРІР‚С™ Р РЋР С“Р В РЎвЂќР В РЎвЂўР РЋР вЂљР В РЎвЂў</span>'
            )
        
        return format_html(
            '<span style="background:#28a745;color:white;padding:6px 12px;border-radius:6px;font-weight:600;">Р В РІР‚в„ўР В Р’В°Р В Р’В»Р В РЎвЂР В РўвЂР В Р’ВµР В Р вЂ¦</span>'
        )
    
    token_status.short_description = "Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“"
    
    def expires_display(self, obj):
        """Р В РІР‚СњР В Р’В°Р РЋРІР‚С™Р В Р’В° Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р В Р’ВµР РЋРІР‚РЋР В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ"""
        if obj.expires_at:
            return obj.expires_at.strftime('%d.%m.%Y %H:%M')
        return "Р Р†Р вЂљРІР‚Сњ"
    
    expires_display.short_description = "Р В Р РЋР С“Р РЋРІР‚С™Р В Р’ВµР В РЎвЂќР В Р’В°Р В Р’ВµР РЋРІР‚С™"
    
    def time_left_display(self, obj):
        """Р В РЎвЂєР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р вЂ Р РЋРІвЂљВ¬Р В Р’ВµР В Р’ВµР РЋР С“Р РЋР РЏ Р В Р вЂ Р РЋР вЂљР В Р’ВµР В РЎР РЋР РЏ"""
        
        if not obj.expires_at:
            return "Р Р†Р вЂљРІР‚Сњ"
        
        time_left = obj.expires_at - timezone.now()
        
        if time_left.total_seconds() < 0:
            return format_html('<span style="color:#dc3545;font-weight:600;">Р В Р РЋР С“Р РЋРІР‚С™Р РЋРІР‚Р В РЎвЂќ</span>')
        
        days = time_left.days
        hours = int(time_left.seconds / 3600)
        minutes = int((time_left.seconds % 3600) / 60)
        
        if time_left.total_seconds() < 3600:
            color = '#dc3545'
        elif time_left.total_seconds() < 7200:
            color = '#ffc107'
        else:
            color = '#28a745'
        
        if days > 0:
            text = f"{days} Р В РўвЂР В Р вЂ¦. {hours} Р РЋРІР‚РЋ."
        elif hours > 0:
            text = f"{hours} Р РЋРІР‚РЋ. {minutes} Р В РЎР В РЎвЂР В Р вЂ¦."
        else:
            text = f"{minutes} Р В РЎР В РЎвЂР В Р вЂ¦."
        
        return format_html('<span style="color:{};font-weight:600;">{}</span>', color, text)
    
    time_left_display.short_description = "Р В РЎвЂєР РЋР С“Р РЋРІР‚С™Р В Р’В°Р В Р’В»Р В РЎвЂўР РЋР С“Р РЋР Р‰"
    
    def action_buttons(self, obj):
        """Р В РЎв„ўР В Р вЂ¦Р В РЎвЂўР В РЎвЂ”Р В РЎвЂќР В РЎвЂ Р В РўвЂР В Р’ВµР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р вЂ Р В РЎвЂР В РІвЂћвЂ“"""
        return format_html('''
            <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <a href="/admin/main/amocrmtoken/refresh/" 
                   class="button" 
                   style="background:#007bff;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;white-space:nowrap;">
                    Р В РЎвЂєР В Р’В±Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р РЋРІР‚С™Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦
                </a>
                <a href="/admin/main/amocrmtoken/logs/" 
                   class="button" 
                   style="background:#dc3545;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;white-space:nowrap;">
                    Р В РІР‚С”Р В РЎвЂўР В РЎвЂ“Р В РЎвЂ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР В РЎвЂќ
                </a>
                <a href="/admin/main/amocrmtoken/instructions/" 
                   class="button" 
                   style="background:#6c757d;color:white;padding:6px 12px;border-radius:4px;text-decoration:none;white-space:nowrap;">
                    Р В Р В Р вЂ¦Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР РЋРЎвЂњР В РЎвЂќР РЋРІР‚В Р В РЎвЂР РЋР РЏ
                </a>
            </div>
        ''')
    
    action_buttons.short_description = "Р В РІР‚СњР В Р’ВµР В РІвЂћвЂ“Р РЋР С“Р РЋРІР‚С™Р В Р вЂ Р В РЎвЂР РЋР РЏ"
    
    # ========== Р В РЎС™Р В РЎвЂ™Р В Р’В Р В Р РѓР В Р’В Р В Р в‚¬Р В РЎС›Р В Р’В« ==========
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('refresh/', self.admin_site.admin_view(self.refresh_token_view), name='amocrm_refresh'),
            path('logs/', self.admin_site.admin_view(self.logs_view), name='amocrm_logs'),
            path('instructions/', self.admin_site.admin_view(self.instructions_view), name='amocrm_instructions'),
        ]
        return custom_urls + urls
    
    # ========== Р В РЎвЂєР В РІР‚Р В Р’В Р В РЎвЂ™Р В РІР‚Р В РЎвЂєР В РЎС›Р В Р’В§Р В Р В РЎв„ўР В  ==========
    def refresh_token_view(self, request):
        """Р В РЎвЂєР В Р’В±Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р РЋРІР‚С™Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦ Р В Р вЂ Р РЋР вЂљР РЋРЎвЂњР РЋРІР‚РЋР В Р вЂ¦Р РЋРЎвЂњР РЋР вЂ№"""
        try:
            token_obj = AmoCRMToken.get_instance()
            
            if not token_obj.refresh_token:
                messages.error(request, 'Refresh token Р В Р вЂ¦Р В Р’Вµ Р В Р вЂ¦Р В Р’В°Р В РІвЂћвЂ“Р В РўвЂР В Р’ВµР В Р вЂ¦. Р В РЎСљР В Р’В°Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В РІвЂћвЂ“Р РЋРІР‚С™Р В Р’Вµ Р РЋРІР‚С™Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦Р РЋРІР‚в„– Р В Р’В·Р В Р’В°Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В РЎвЂў.')
                return redirect('/admin/main/amocrmtoken/')
            
            TokenManager.refresh_token(token_obj)
            messages.success(request, f'Р В РЎС›Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦ Р РЋРЎвЂњР РЋР С“Р В РЎвЂ”Р В Р’ВµР РЋРІвЂљВ¬Р В Р вЂ¦Р В РЎвЂў Р В РЎвЂўР В Р’В±Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р РЋРІР‚Р В Р вЂ¦. Р В Р РЋР С“Р РЋРІР‚С™Р В Р’ВµР В РЎвЂќР В Р’В°Р В Р’ВµР РЋРІР‚С™: {token_obj.expires_at.strftime("%d.%m.%Y %H:%M")}')
            
        except Exception as e:
            messages.error(request, f'Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р В РЎвЂўР В Р’В±Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ: {str(e)}')
        
        return redirect('/admin/main/amocrmtoken/')
    
    def logs_view(self, request):
        """Р В РЎСџР В РЎвЂўР В РЎвЂќР В Р’В°Р В Р’В·Р В Р’В°Р РЋРІР‚С™Р РЋР Р‰ Р В Р’В»Р В РЎвЂўР В РЎвЂ“Р В РЎвЂ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР В РЎвЂќ amoCRM"""

        
        amocrm_log_path = os.path.join(settings.BASE_DIR, 'logs', 'amocrm.log')
        errors_log_path = os.path.join(settings.BASE_DIR, 'logs', 'errors.log')
        
        amocrm_logs = []
        errors_logs = []
        
        if os.path.exists(amocrm_log_path):
            try:
                with open(amocrm_log_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    amocrm_logs = [line.strip() for line in lines if line.strip()][-100:]
            except Exception as e:
                amocrm_logs = [f"Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р РЋРІР‚РЋР РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ amocrm.log: {str(e)}"]
        
        if os.path.exists(errors_log_path):
            try:
                with open(errors_log_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    errors_logs = [line.strip() for line in lines if line.strip()][-50:]
            except Exception as e:
                errors_logs = [f"Р В РЎвЂєР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂќР В Р’В° Р РЋРІР‚РЋР РЋРІР‚С™Р В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋР РЏ errors.log: {str(e)}"]
        
        context = {
            **self.admin_site.each_context(request),
            'title': 'Р В РІР‚С”Р В РЎвЂўР В РЎвЂ“Р В РЎвЂ Р В РЎвЂўР РЋРІвЂљВ¬Р В РЎвЂР В Р’В±Р В РЎвЂўР В РЎвЂќ amoCRM',
            'amocrm_logs': amocrm_logs,
            'errors_logs': errors_logs,
        }
        return render(request, 'main/amocrm_logs.html', context)
    
    def instructions_view(self, request):
        """Р В РЎСџР В РЎвЂўР В РЎвЂќР В Р’В°Р В Р’В·Р В Р’В°Р РЋРІР‚С™Р РЋР Р‰ Р В РЎвЂР В Р вЂ¦Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР РЋРЎвЂњР В РЎвЂќР РЋРІР‚В Р В РЎвЂР РЋР вЂ№"""
        
        token_obj = AmoCRMToken.get_instance()
        time_left_text = None
        
        if token_obj.expires_at:
            time_left = token_obj.expires_at - timezone.now()
            
            if time_left.total_seconds() < 0:
                time_left_text = "Р В РЎС›Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦ Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р РЋРІР‚Р В РЎвЂќ"
            else:
                days = time_left.days
                hours = int(time_left.seconds / 3600)
                minutes = int((time_left.seconds % 3600) / 60)
                
                parts = []
                if days > 0:
                    parts.append(f"{days} Р В РўвЂР В Р вЂ¦.")
                if hours > 0:
                    parts.append(f"{hours} Р РЋРІР‚РЋ.")
                if minutes > 0:
                    parts.append(f"{minutes} Р В РЎР В РЎвЂР В Р вЂ¦.")
                
                time_left_text = " ".join(parts) if parts else "Р В РЎС™Р В Р’ВµР В Р вЂ¦Р В Р’ВµР В Р’Вµ Р В РЎР В РЎвЂР В Р вЂ¦Р РЋРЎвЂњР РЋРІР‚С™Р РЋРІР‚в„–"
        
        context = {
            **self.admin_site.each_context(request),
            'title': 'Р В Р В Р вЂ¦Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР РЋРЎвЂњР В РЎвЂќР РЋРІР‚В Р В РЎвЂР РЋР РЏ: Р В РЎвЂєР В Р’В±Р В Р вЂ¦Р В РЎвЂўР В Р вЂ Р В Р’В»Р В Р’ВµР В Р вЂ¦Р В РЎвЂР В Р’Вµ Р РЋРІР‚С™Р В РЎвЂўР В РЎвЂќР В Р’ВµР В Р вЂ¦Р В РЎвЂўР В Р вЂ  amoCRM',
            'token': token_obj,
            'time_left_text': time_left_text,
        }
        return render(request, 'main/amocrm_instructions.html', context)
    
# ========== DILERLAR ==========

class DealerImageInline(admin.TabularInline):
    model = DealerImage
    extra = 1
    fields = ('image', 'caption_uz', 'caption_ru', 'caption_en', 'order')
    ordering = ('order',)


@admin.register(Dealer)
class DealerAdmin(ContentAdminMixin, TabbedTranslationAdmin):
    list_display = ('name', 'region', 'phone', 'is_active', 'order')
    list_filter = ('region', 'is_active')
    list_editable = ('is_active', 'order')
    search_fields = ('name', 'address', 'phone')
    ordering = ('order', 'name')
    inlines = [DealerImageInline]
    fieldsets = (
        ("Asosiy ma'lumotlar", {
            'fields': (
                'name_uz', 'name_ru', 'name_en',
                'region',
                'address_uz', 'address_ru', 'address_en',
                'phone',
                'working_hours_uz', 'working_hours_ru', 'working_hours_en',
                'logo',
            )
        }),
        ("Tavsif", {
            'fields': ('description_uz', 'description_ru', 'description_en'),
            'classes': ('collapse',),
        }),
        ("Xarita koordinatalari", {
            'fields': ('latitude', 'longitude'),
        }),
        ("Ijtimoiy tarmoqlar", {
            'fields': ('instagram', 'telegram', 'facebook', 'youtube'),
            'classes': ('collapse',),
        }),
        ("Sozlamalar", {
            'fields': ('is_active', 'order'),
        }),
    )

    class Media:
        css = {'all': ('css/custom_dealer_admin.css',)}
        js = ('js/admin/dealer_hide_lang_select.js',)


# ========== Р В РЎвЂєР В РЎС›Р В РІР‚вЂќР В Р’В«Р В РІР‚в„ўР В Р’В« Р В РЎв„ўР В РІР‚С”Р В Р В РІР‚СћР В РЎСљР В РЎС›Р В РЎвЂєР В РІР‚в„ў ==========

@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ('name', 'rating_stars', 'status_badge', 'is_verified', 'created_at', 'moderated_at')
    list_filter = ('status', 'rating', 'is_verified', 'created_at')
    list_editable = ('is_verified',)
    search_fields = ('name', 'text')
    readonly_fields = ('ip_address', 'created_at', 'moderated_at', 'avatar_preview')
    ordering = ('-created_at',)
    actions = ['approve_reviews', 'reject_reviews']
    list_per_page = 30

    fieldsets = (
        ("Р В Р В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ Р В РЎвЂўР В Р’В± Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р В Р’Вµ", {
            'fields': ('name', 'rating', 'text', 'avatar', 'avatar_preview')
        }),
        ("Р В РЎС™Р В РЎвЂўР В РўвЂР В Р’ВµР РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ", {
            'fields': ('status', 'is_verified', 'admin_comment'),
            'description': 'Р Р†РЎв„ўР’В Р С—РЎвЂР РЏ Р В РЎС›Р В РЎвЂўР В Р’В»Р РЋР Р‰Р В РЎвЂќР В РЎвЂў Р В РЎвЂўР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’ВµР В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р РЋРІР‚в„– Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’В°Р В Р’В¶Р В Р’В°Р РЋР вЂ№Р РЋРІР‚С™Р РЋР С“Р РЋР РЏ Р В Р вЂ¦Р В Р’В° Р РЋР С“Р В Р’В°Р В РІвЂћвЂ“Р РЋРІР‚С™Р В Р’Вµ'
        }),
        ("Р В Р Р‹Р В РЎвЂР РЋР С“Р РЋРІР‚С™Р В Р’ВµР В РЎР В Р вЂ¦Р В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ", {
            'fields': ('ip_address', 'created_at', 'moderated_at'),
            'classes': ('collapse',),
        }),
    )

    def rating_stars(self, obj):
        return format_html(
            '<span style="color:#F7941D;font-size:16px;">{}</span>',
            'Р Р†РІР‚В¦' * obj.rating + 'Р Р†РІР‚В ' * (5 - obj.rating)
        )
    rating_stars.short_description = "Р В РЎвЂєР РЋРІР‚В Р В Р’ВµР В Р вЂ¦Р В РЎвЂќР В Р’В°"

    def status_badge(self, obj):
        colors = {'pending': '#FF9800', 'approved': '#4CAF50', 'rejected': '#f44336'}
        labels = {'pending': 'Р В РЎСљР В Р’В° Р В РЎР В РЎвЂўР В РўвЂР В Р’ВµР РЋР вЂљР В Р’В°Р РЋРІР‚В Р В РЎвЂР В РЎвЂ', 'approved': 'Р В РЎвЂєР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’ВµР В Р вЂ¦', 'rejected': 'Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР В Р’В»Р В РЎвЂўР В Р вЂ¦Р РЋРІР‚Р В Р вЂ¦'}
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;border-radius:12px;font-size:12px;">{}</span>',
            colors.get(obj.status, '#999'),
            labels.get(obj.status, obj.status)
        )
    status_badge.short_description = "Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“"

    def avatar_preview(self, obj):
        if obj.avatar:
            return format_html('<img src="{}" style="max-width:150px;max-height:150px;border-radius:50%;">', obj.avatar.url)
        return "Р В РЎСљР В Р’ВµР РЋРІР‚С™ Р РЋРІР‚С›Р В РЎвЂўР РЋРІР‚С™Р В РЎвЂў"
    avatar_preview.short_description = "Р В РЎСџР РЋР вЂљР В Р’ВµР В Р вЂ Р РЋР Р‰Р РЋР вЂ№ Р В Р’В°Р В Р вЂ Р В Р’В°Р РЋРІР‚С™Р В Р’В°Р РЋР вЂљР В Р’В°"

    def approve_reviews(self, request, queryset):
        count = queryset.update(status='approved', is_verified=True, moderated_at=timezone.now())
        self.message_user(request, f"Р Р†РЎС™РІР‚В¦ Р В РЎвЂєР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В Р’ВµР В Р вЂ¦Р В РЎвЂў Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р В РЎвЂўР В Р вЂ : {count}", messages.SUCCESS)
    approve_reviews.short_description = "Р Р†РЎС™РІР‚В¦ Р В РЎвЂєР В РўвЂР В РЎвЂўР В Р’В±Р РЋР вЂљР В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р РЋРІР‚в„–"

    def reject_reviews(self, request, queryset):
        count = queryset.update(status='rejected', moderated_at=timezone.now())
        self.message_user(request, f"Р Р†РЎСљР Р‰ Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР В Р’В»Р В РЎвЂўР В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р В РЎвЂў Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р В РЎвЂўР В Р вЂ : {count}", messages.WARNING)
    reject_reviews.short_description = "Р Р†РЎСљР Р‰ Р В РЎвЂєР РЋРІР‚С™Р В РЎвЂќР В Р’В»Р В РЎвЂўР В Р вЂ¦Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ Р В РЎвЂўР РЋРІР‚С™Р В Р’В·Р РЋРІР‚в„–Р В Р вЂ Р РЋРІР‚в„–"

    def save_model(self, request, obj, form, change):
        if change and 'status' in form.changed_data:
            obj.moderated_at = timezone.now()
        super().save_model(request, obj, form, change)


# ========== Р В РЎС›Р В РІР‚СћР В Р Р‹Р В РЎС›-Р В РІР‚СњР В Р’В Р В РЎвЂ™Р В РІвЂћСћР В РІР‚в„ў ==========

@admin.register(TestDriveRequest)
class TestDriveRequestAdmin(admin.ModelAdmin):
    list_display = ['name', 'phone', 'product', 'dealer', 'preferred_date', 'preferred_time', 'status', 'created_at']
    list_filter = ['status', 'dealer', 'preferred_date', 'created_at']
    search_fields = ['name', 'phone']
    list_editable = ['status']
    readonly_fields = ['ip_address', 'referer', 'utm_data', 'visitor_uid', 'created_at', 'updated_at']
    ordering = ['-created_at']

    fieldsets = (
        ('Р В РЎв„ўР В Р’В»Р В РЎвЂР В Р’ВµР В Р вЂ¦Р РЋРІР‚С™', {
            'fields': ('name', 'phone', 'dealer', 'product')
        }),
        ('Р В РІР‚СњР В Р’В°Р РЋРІР‚С™Р В Р’В° Р В РЎвЂ Р В Р вЂ Р РЋР вЂљР В Р’ВµР В РЎР РЋР РЏ', {
            'fields': ('preferred_date', 'preferred_time')
        }),
        ('Р В Р Р‹Р РЋРІР‚С™Р В Р’В°Р РЋРІР‚С™Р РЋРЎвЂњР РЋР С“', {
            'fields': ('status', 'admin_comment')
        }),
        ('Р В РЎС›Р В Р’ВµР РЋРІР‚В¦Р В Р вЂ¦Р В РЎвЂР РЋРІР‚РЋР В Р’ВµР РЋР С“Р В РЎвЂќР В Р’В°Р РЋР РЏ Р В РЎвЂР В Р вЂ¦Р РЋРІР‚С›Р В РЎвЂўР РЋР вЂљР В РЎР В Р’В°Р РЋРІР‚В Р В РЎвЂР РЋР РЏ', {
            'classes': ('collapse',),
            'fields': ('ip_address', 'referer', 'utm_data', 'visitor_uid', 'created_at', 'updated_at')
        }),
    )

    actions = ['mark_confirmed', 'mark_completed', 'mark_cancelled']

    def mark_confirmed(self, request, queryset):
        count = queryset.update(status='confirmed')
        self.message_user(request, f"Р В РЎСџР В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В Р’В¶Р В РўвЂР В Р’ВµР В Р вЂ¦Р В РЎвЂў: {count}", messages.SUCCESS)
    mark_confirmed.short_description = "Р В РЎСџР В РЎвЂўР В РўвЂР РЋРІР‚С™Р В Р вЂ Р В Р’ВµР РЋР вЂљР В РўвЂР В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ"

    def mark_completed(self, request, queryset):
        count = queryset.update(status='completed')
        self.message_user(request, f"Р В РІР‚вЂќР В Р’В°Р В Р вЂ Р В Р’ВµР РЋР вЂљР РЋРІвЂљВ¬Р В Р’ВµР В Р вЂ¦Р В РЎвЂў: {count}", messages.SUCCESS)
    mark_completed.short_description = "Р В РЎвЂєР РЋРІР‚С™Р В РЎР В Р’ВµР РЋРІР‚С™Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В РЎвЂќР В Р’В°Р В РЎвЂќ Р В Р’В·Р В Р’В°Р В Р вЂ Р В Р’ВµР РЋР вЂљР РЋРІвЂљВ¬Р РЋРІР‚Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ"

    def mark_cancelled(self, request, queryset):
        count = queryset.update(status='cancelled')
        self.message_user(request, f"Р В РЎвЂєР РЋРІР‚С™Р В РЎР В Р’ВµР В Р вЂ¦Р В Р’ВµР В Р вЂ¦Р В РЎвЂў: {count}", messages.WARNING)
    mark_cancelled.short_description = "Р В РЎвЂєР РЋРІР‚С™Р В РЎР В Р’ВµР В Р вЂ¦Р В РЎвЂР РЋРІР‚С™Р РЋР Р‰ Р В Р вЂ Р РЋРІР‚в„–Р В Р’В±Р РЋР вЂљР В Р’В°Р В Р вЂ¦Р В Р вЂ¦Р РЋРІР‚в„–Р В Р’Вµ"


# ========== Р В РЎв„ўР В РЎвЂєР В РЎС™Р В РЎвЂ™Р В РЎСљР В РІР‚СњР В РЎвЂ™ Р Р†Р вЂљРІР‚Сњ Р В РЎС™Р В РІР‚СћР В РЎСљР В РІР‚СћР В РІР‚СњР В РІР‚вЂњР В РІР‚СћР В Р’В Р В Р’В« ==========

@admin.register(BranchManager)
class BranchManagerAdmin(TabbedTranslationAdmin):
    list_display = ['full_name', 'position', 'dealer', 'phone', 'is_active', 'order']
    list_filter = ['dealer', 'is_active']
    list_editable = ['is_active', 'order']
    search_fields = ['full_name', 'phone']
    ordering = ['dealer__order', 'order']

    fieldsets = (
        (None, {
            'fields': ('full_name', 'position', 'photo', 'dealer', 'phone')
        }),
        ('Р В РЎСљР В Р’В°Р РЋР С“Р РЋРІР‚С™Р РЋР вЂљР В РЎвЂўР В РІвЂћвЂ“Р В РЎвЂќР В РЎвЂ', {
            'fields': ('is_active', 'order')
        }),
    )


@admin.register(TelegramUser)
class TelegramUserAdmin(admin.ModelAdmin):
    list_display = ['telegram_id', 'username', 'first_name', 'phone', 'region', 'language', 'age', 'created_at']
    list_filter = ['region', 'language', 'created_at']
    search_fields = ['username', 'first_name', 'phone', 'telegram_id']
    readonly_fields = ['telegram_id', 'username', 'first_name', 'phone', 'region', 'language', 'age', 'created_at']
    ordering = ['-created_at']
    list_per_page = 50





    